"""
Excel 写入器 — 将解析后的数据输出为 Excel 文件。

支持两种模式:
    1. 合并模式: 多个 TXT → 单一 .xlsx（每个 TXT 一个 Sheet）
    2. 独立模式: 每个 TXT → 独立的 .xlsx 文件
"""

import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# openpyxl Sheet 名称限制: 最长 31 个字符，不能包含 \ / * ? : [ ]
_SHEET_NAME_ILLEGAL = re.compile(r'[\\/*?:\[\]]')
_MAX_SHEET_NAME_LEN = 31


def sanitize_sheet_name(filename):
    """
    将文件名转换为合法的 Excel Sheet 名称。

    Args:
        filename: 文件名（可能含扩展名）

    Returns:
        str: 合法的 Sheet 名称（≤31 字符，无非法字符）
    """
    # 去掉扩展名
    name = os.path.splitext(os.path.basename(filename))[0]
    # 替换非法字符
    name = _SHEET_NAME_ILLEGAL.sub("_", name)
    # 截断到 31 字符
    if len(name) > _MAX_SHEET_NAME_LEN:
        name = name[:_MAX_SHEET_NAME_LEN]
    # 不能为空
    if not name:
        name = "Sheet"
    return name


def _resolve_sheet_name_conflicts(sheet_names):
    """
    解决 Sheet 名称冲突：为重复的名称添加 _1, _2 后缀。

    Args:
        sheet_names: 原始 Sheet 名称列表

    Returns:
        list[str]: 去重后的 Sheet 名称列表
    """
    seen = {}
    result = []
    for name in sheet_names:
        if name in seen:
            seen[name] += 1
            # 生成不冲突的新名称
            suffix = f"_{seen[name]}"
            max_base = _MAX_SHEET_NAME_LEN - len(suffix)
            candidate = name[:max_base] + suffix
            # 极端情况：截断后仍冲突则加数字
            while candidate in seen:
                seen[name] += 1
                suffix = f"_{seen[name]}"
                max_base = _MAX_SHEET_NAME_LEN - len(suffix)
                candidate = name[:max_base] + suffix
            result.append(candidate)
            seen[candidate] = 0  # 标记已使用
        else:
            seen[name] = 0
            result.append(name)
    return result


def _auto_column_widths(ws, channels, rows, min_width=10, max_width=20):
    """
    根据数据内容直接计算并设置列宽。

    Args:
        ws: openpyxl Worksheet
        channels: 通道名列表
        rows: 数据行列表
        min_width: 最小列宽
        max_width: 最大列宽
    """
    num_cols = len(channels)
    # 取前 500 行作为采样（足够判断宽度，避免全量遍历）
    sample_rows = rows[:500]

    for col_idx in range(num_cols):
        # 从表头开始
        max_len = len(str(channels[col_idx]))
        # 检查采样数据行
        for row_data in sample_rows:
            if col_idx < len(row_data):
                val = row_data[col_idx]
                if val is not None:
                    max_len = max(max_len, len(str(val)))

        # 列宽：字符宽度 + 2 的余量，限制在合理范围
        width = max(min_width, min(max_len + 3, max_width))
        col_letter = get_column_letter(col_idx + 1)
        ws.column_dimensions[col_letter].width = width


def _write_sheet(ws, channels, rows, freeze_header=True):
    """
    将通道和数据写入工作表的通用函数。

    性能说明: 使用 ws.append() 批量写入整行，避免逐单元格调用
    ws.cell(row, col, value=...) 造成的 O(n*m) 次对象创建开销。
    对 10000+ 行的典型文件，速度提升 10-50 倍。

    Args:
        ws: openpyxl Worksheet
        channels: 通道名列表（用作表头）
        rows: 数据行列表
        freeze_header: 是否冻结首行
    """
    # 样式定义
    header_font = Font(name="Times New Roman", bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="Times New Roman", size=11)

    # 写入表头（append 一次性写入整行）
    ws.append(channels)
    for col_idx in range(1, len(channels) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 批量写入数据行：append() 是 openpyxl 最快的逐行写入方式
    # 数据为纯数字（32位无符号整数）
    for row_data in rows:
        ws.append(row_data)

    # 设置数据区字体（Times New Roman）
    for row in ws.iter_rows(min_row=2, max_row=len(rows) + 1,
                            min_col=1, max_col=len(channels)):
        for cell in row:
            cell.font = data_font

    # 自动调整列宽（基于数据直接计算）
    _auto_column_widths(ws, channels, rows)

    # 冻结首行
    if freeze_header:
        ws.freeze_panes = "A2"


def write_single_workbook(parsed_files, output_path, unified_channels=None):
    """
    合并模式：将所有文件写入同一个 Excel 工作簿的不同 Sheet。

    每个 Sheet 保留各自文件的通道列，不强制统一为超集。
    这样每个文件只显示自己实际拥有的通道，不会多出全空列。

    Args:
        parsed_files: parser.parse_txt_file() 返回结果的列表
        output_path: 输出 .xlsx 文件的完整路径
        unified_channels: 已弃用，保留仅为兼容旧调用

    Returns:
        tuple: (success: bool, message: str)
    """
    if not parsed_files:
        return False, "没有可写入的数据"

    wb = Workbook()

    try:
        # 生成 Sheet 名称并解决冲突
        raw_names = [sanitize_sheet_name(pf["filename"]) for pf in parsed_files]
        sheet_names = _resolve_sheet_name_conflicts(raw_names)

        for i, pf in enumerate(parsed_files):
            if i == 0:
                ws = wb.active
                ws.title = sheet_names[i]
            else:
                ws = wb.create_sheet(title=sheet_names[i])

            # 每个 Sheet 使用文件自己的通道，不强制统一
            _write_sheet(ws, pf["channels"], pf["rows"])

        wb.save(output_path)
        return True, f"成功生成: {output_path}（共 {len(parsed_files)} 个 Sheet）"
    except (IOError, OSError) as e:
        return False, f"保存文件失败: {e}"
    except Exception as e:
        return False, f"写入 Excel 时出错: {e}"


def write_individual_files(parsed_files, output_dir):
    """
    独立模式：每个文件生成一个独立的 .xlsx 文件。

    Args:
        parsed_files: parser.parse_txt_file() 返回结果的列表
        output_dir: 输出目录路径

    Returns:
        list[tuple]: [(filename, success: bool, message: str), ...]
    """
    if not parsed_files:
        return [("", False, "没有可写入的数据")]

    results = []
    for pf in parsed_files:
        base_name = os.path.splitext(pf["filename"])[0]
        output_path = os.path.join(output_dir, f"{base_name}.xlsx")

        # 处理文件名冲突：如果已存在则加序号
        counter = 1
        while os.path.exists(output_path):
            output_path = os.path.join(output_dir, f"{base_name}_{counter}.xlsx")
            counter += 1

        wb = Workbook()
        try:
            ws = wb.active
            ws.title = sanitize_sheet_name(pf["filename"])
            _write_sheet(ws, pf["channels"], pf["rows"])
            wb.save(output_path)
            results.append((pf["filename"], True, output_path))
        except (IOError, OSError) as e:
            results.append((pf["filename"], False, f"保存失败: {e}"))
        except Exception as e:
            results.append((pf["filename"], False, f"写入出错: {e}"))

    return results
